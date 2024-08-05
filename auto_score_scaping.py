from selenium import webdriver
import selenium
from selenium.webdriver.common.by import By
import cv2
import pytesseract
import openpyxl
import argparse
import os
import io
import numpy as np
from PIL import Image
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


url = "https://ap.ceec.edu.tw/RegExam/ScoreSearch/Login?examtype=B"

# ID = "21035415"
# PID = "L125715002"



# cool_cookie = {'name': 'TS011ec7b5', 'value': '01e5022d03271e1ae0c88d2971b19ce0b143d450b93769256ff61234394910a9d6d95478c412faece818d94da1265e8544eea522aa5a0da5a6b24114ce9c70115ee317edfbb637a0842ee5249f70e267a5b298fe5a9498db67cdbcc231baa857b0ccdb90930be70e43beedb66d3a0cbd55bd7d45a087b18eb58a4dfb46f2e28ca2b7eb23970d48294c47535e3f2d8265c20b3f1fd50f0c1d1670045f97dd906fc7a28aa6f2', 'path': '/', 'domain': '.ap.ceec.edu.tw', 'secure': False, 'httpOnly': False, 'sameSite': 'None'}
# TS011ec7b5 = {'name': 'TS011ec7b5', 'value': '01e5022d0319d8c328a2e7dda65ca8780791e1dc4c7cb22a9b2615103c191e5a3ff64378acc0b50cf47c387953879535558eb5b6e376094bf5db0b650dc2d49623fa2d1a77a213f4134520ab0f3a8c8d20a04fad1205486450f0deb8895c7a6a253582d05dce03b81a3d668acdb2a49a22694f4821971bc8c24b08ae2c0fdbee1d848114e6af4c7558d89f956135f514ffd25d5b46bbce1f105157f2ce043a28d682938260', 'path': '/', 'domain': '.ap.ceec.edu.tw', 'secure': False, 'httpOnly': False, 'sameSite': 'None'}
# ADRUM_BTa = {'name': 'ADRUM_BTa', 'value': 'R:23|g:9b3f18d7-ad5e-404b-bae0-f9b923c0e4fc|n:customer1_50aad0b3-5dcf-47b6-a3eb-222b98a41148', 'path': '/', 'domain': 'ap.ceec.edu.tw', 'secure': True, 'httpOnly': True, 'expiry': 1722232079, 'sameSite': 'None'}

# input("=====================================")

def solve_captcha(driver):
    png = driver.get_screenshot_as_png()
    screenshot = Image.open(io.BytesIO(png))
    screenshot = np.array(screenshot)
    # image = image[loc['y']: loc['y']+500, loc['x']: loc['x']+500]
    image = screenshot[1170: 1216, 809: 959]
    raw_string = pytesseract.image_to_string(image, lang="eng", config='--psm 13 -c tessedit_char_whitelist=0123456789').replace("\n", '')
    print(raw_string)
    btn = driver.find_element(By.XPATH, "/html/body/div[2]/div/div[1]/div/form/div[3]/button[1]")
    if len(raw_string) != 4 and len(raw_string) != 3:
        btn.click()
        solve_captcha(driver)
        return
    
    solution = int(raw_string[:2]) + int(raw_string[-1])  
    print(solution)
    elem = driver.find_element(By.ID, "Captcha")
    elem.send_keys(solution)
    
def login(driver, ID, PID, fail_count=0):
    if fail_count > 5:
        return -1
    driver.get(url)
    elem = driver.find_element(By.ID, "TestID")
    elem.send_keys(ID)
    elem = driver.find_element(By.ID, "PID")
    elem.send_keys(PID)
    solve_captcha(driver)
    driver.find_element(By.ID, "login").click()

    # img = driver.find_element(By.ID, "valiCode")
    # loc = img.location
    
    # input("wait")
    if driver.current_url != "https://ap.ceec.edu.tw/RegExam/ScoreSearch/StuResult":
        # prompt_text = driver.find_element(By.XPATH, "/html/body/div[4]/div[2]/div/div/div/div/div/div/div/div[3]/div/div").text
        # while (not prompt_text):
        
        #     prompt_text = driver.find_element(By.XPATH, "/html/body/div[4]/div[2]/div/div/div/div/div/div/div/div[3]/div/div").text
        # if prompt_text == "成績查詢已保密":
        #     print("成績查詢已保密")
        #     return "secret"
        
        wait = WebDriverWait(driver, 5)
        while(True):
            try:
                prompt_text = wait.until(
                    # EC.presence_of_element_located((By.CLASS_NAME, "jconfirm-title"))
                    # EC.text_to_be_present_in_element_value((By.CLASS_NAME, 'jconfirm-title'), "訊息")
                    EC.presence_of_element_located((By.XPATH, "/html/body/div[4]/div[2]/div/div/div/div/div/div/div/div[3]/div/div"))
                    # EC.visibility_of_element_located((By.XPATH, "/html/body/div[4]/div[2]/div/div/div/div/div/div/div/div[3]/div/div"))
                    ).text
            except selenium.common.exceptions.TimeoutException:
                return
            if prompt_text: break
            
        print(f'"{prompt_text}"')
        if prompt_text == "成績查詢已保密":
            print(prompt_text)
            return "secret"
        return login(driver, ID, PID, fail_count+1)

# out = cv2.CreateImage((150,60), image.depth, 3)
# cv2.SetImageROI(image, (loc['x'],loc['y'],150,60))
# cv2.Resize(image, out)
# cv2.imshow("processed image", image)
# cv2.waitKey(0)
# input("wait")
# print(driver.get_cookie("TS011ec7b5"))
# print("=====================================")
# print(driver.get_cookie("ADRUM_BTa"))

# driver.delete_cookie("TS011ec7b5")
# driver.add_cookie(TS011ec7b5)
# driver.delete_cookie("ADRUM_BTa")
# driver.add_cookie(ADRUM_BTa)

def get_data_ast(driver):
    # try:
        # input('wait')
    data = {
    "math_alpha": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[3]/div/table/tbody/tr/td[1]").text,
    "chemistry": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[3]/div/table/tbody/tr/td[2]").text,
    "physics": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[3]/div/table/tbody/tr/td[3]").text,
    "biology": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[3]/div/table/tbody/tr/td[4]").text,
    "history": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[3]/div/table/tbody/tr/td[5]").text,
    "geography": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[3]/div/table/tbody/tr/td[6]").text,
    "citizen": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[3]/div/table/tbody/tr/td[7]").text,
    "chinese": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[5]/div/table/tbody/tr/td[1]").text,
    "english": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[5]/div/table/tbody/tr/td[2]").text,
    "mathA": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[5]/div/table/tbody/tr/td[3]").text,
    "mathB": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[5]/div/table/tbody/tr/td[4]").text,
    "society": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[5]/div/table/tbody/tr/td[5]").text,
    "science": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[5]/div/table/tbody/tr/td[6]").text,
    }
    # except selenium.common.exceptions.NoSuchElementException: return -1
    return data
def get_data_gsat(driver):
    try:
        data = {
        "chinese": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[3]/div[2]").text,
        "english": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[4]/div[2]").text,
        "mathA": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[5]/div[2]").text,
        "mathB": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[6]/div[2]").text,
        "society": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[7]/div[2]").text,
        "society": driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div/div[8]/div[2]").text,
        }
    except selenium.common.exceptions.NoSuchElementException: return -1
    return data
def set_header_gsat(sheet_obj, start, output_path):
    objects_list = ['國文', '英文', '數學A', '數學B', '社會', '自然']
    for subject, i in zip(objects_list, range(13)):
        sheet_obj.cell(row=1, column=start+i).value = subject
    wb_obj.save(output_path)
    
def save_data_gsat(data, sheet_obj, start, row, output_path):
    subject_list = ["chin", "english", "mathA", "mathB", "society", "science"]
    for subject, i in zip(subject_list, range(0, len(subject_list)+1)):
        sheet_obj.cell(row=row, column=start+i).value = data[subject]
    wb_obj.save(output_path)

def set_header(sheet_obj, start, output_path):
    objects_list = ['數學甲', '化學', '物理', '生物', '歷史', '地理', '公民', '國文', '英文', '數學Ａ', '數學Ｂ', '社會', '自然']
    for subject, i in zip(objects_list, range(13)):
        sheet_obj.cell(row=1, column=start+i).value = subject
    wb_obj.save(output_path)
    
def save_data(data, sheet_obj, start, row, output_path):
    print(data)
    subject_list = ["math_alpha", "chemistry", "physics", "biology", "history", "geography", "citizen", "chinese", "english", "mathA", "mathB", "society", "science"]
    for subject, i in zip(subject_list, range(0, len(subject_list)+1)):
        sheet_obj.cell(row=row, column=start+i).value = data[subject]
    wb_obj.save(output_path)

def find_start(sheet_obj, row):
    pointer = 16
    is_empty = False
    while (is_empty == False):
        pointer += 1
        val = str(sheet_obj.cell(row=row, column=pointer).value)
        if val == "數學甲": return pointer
        # if val.isnumeric(): 
        #     val = int(val)
        #     if 0 <= val <=60:
        #         return -1

        if (str(val) =='None'):
            next_val = str(sheet_obj.cell(row=row, column=(pointer+1)).value)
            if (next_val == "None"): 
                return pointer
def find_name(name_list, sheet_obj, row):
    pointer = 1
    while (True):
        pointer += 1
        val = str(sheet_obj.cell(row=row, column=pointer).value)
        if all([i in val for i in name_list]): return pointer

if __name__ == "__main__":
    files = os.listdir('.')
    for file in files:
        if file.startswith("processing_"):
            excel_file_name = file

            print(f"There is an existing {excel_file_name}.")
            if input("Do you want to use this sheet? (y/n)") == "y":
                wb_obj = openpyxl.load_workbook(excel_file_name)
                sheet_obj = wb_obj.active
                output_path = excel_file_name
                # mode = "continue"
                break

    else:
        parser = argparse.ArgumentParser(description="This is a description of your program.")
        
        parser.add_argument("excel_file_name", help="excel原始檔")
        parser.add_argument("--sheet", help="sheet名稱")
        excel_file_name = parser.parse_args().excel_file_name
        excel_sheet_name = parser.parse_args().sheet
        wb_obj = openpyxl.load_workbook(excel_file_name)
        if excel_sheet_name:
            sheet_obj = wb_obj[excel_sheet_name]
        else: 
            sheet_obj = wb_obj.active
        excel_file_name = excel_file_name.split("/")[-1]
        output_path = f"processing_{excel_file_name}"
    # parser.add_argument("--output", required=True, help="output excel檔")
#   wait for altering
    # excels_path = "excels/黎明(方濟會)分科準0728.xlsx"
    # excels_path = "excels/0625港明分科准xlsx.xlsx"
    # excels_path = "excels/0703文華分科准.xlsx"

    # output_path = "processing.xlsx"
    
    # excels_path = "output.xlsx"

    row = sheet_obj.max_row
    column = sheet_obj.max_column

    driver = webdriver.Firefox()
    start = find_start(sheet_obj, 1)
    set_header(sheet_obj, start, output_path)
    col_ID = find_name(['分','准'], sheet_obj, 1)
    col_PID = find_name(["身分證"], sheet_obj, 1)
    for i in range(2, row+1):
        # print(f'"{sheet_obj.cell(row=i, column=start).value}"')
        if str(sheet_obj.cell(row=i, column=start).value) != "None": continue
        ID = str(sheet_obj.cell(row=i, column=col_ID).value).replace("\n", "")
        PID = str(sheet_obj.cell(row=i, column=col_PID).value).replace("\n", "")

        if len(ID) != 8: continue
        print(ID, PID)
        success = login(driver, ID, PID)
        # print(success)
        if success == "secret":
            sheet_obj.cell(row=i, column=start).value = "成績查詢已保密"
            wb_obj.save(output_path)
            continue
        data = get_data_ast(driver)
        # print(data)
        if data == -1:
            continue
        save_data(data, sheet_obj, start, i, output_path)
    wb_obj.save(os.path.join("finished", output_path.split("sing_")[-1] + "(已完成).xlsx"))
    os.remove(output_path)

    # for root, dirs, files in os.walk('.'):
    #     if "processing.xlsx" in files:
    #         os.remove("processing.xlsx")

    driver.quit()
    print("finished")
