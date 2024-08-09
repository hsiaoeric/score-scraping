from selenium import webdriver
import selenium
from selenium.webdriver.common.by import By
import cv2
import pytesseract
import openpyxl
import argparse
import os
from matplotlib import pyplot as plt
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from PIL import Image
import io
import time
import numpy as np

url_B = "https://ap.ceec.edu.tw/RegExam/RegInfo/Login?examtype=B"
url_A = "https://ap.ceec.edu.tw/RegExam/RegInfo/Login?examtype=A"
url = url_A

def solve_captcha(driver):
    # captcha_image = driver.find_element(By.XPATH, "/html/body/div[2]/div/div[1]/div/form/div[3]/button[1]")
    # driver.execute_script("arguments[0].scrollIntoView();", captcha_image)
    # location = image_element.location
    # driver.save_screenshot('screenshot.png')
    # image = cv2.imread('screenshot.png')
    # captcha_image = driver.find_element(By.ID, "valiCode")
    # captcha_image = driver.find_element(By.XPATH, "/html/body/div[2]/div/div/div/form/div[3]/button[2]/i")

    # loc = captcha_image.location
    # print(loc)
    png = driver.get_screenshot_as_png()
    screenshot = Image.open(io.BytesIO(png))
    screenshot = np.array(screenshot)
    # plt.imshow(screenshot)
    # plt.show()
    # image = image[loc['y']: loc['y']+500, loc['x']: loc['x']+500]
    # 800 1174 965 1235
    image = screenshot[1174: 1235, 800: 965]
    # plt.imshow(image)
    # plt.show()
    # input("wait")
    raw_string = pytesseract.image_to_string(image, lang="eng", config='--psm 13 -c tessedit_char_whitelist=0123456789').replace("\n", '')
    print(raw_string)
    btn = driver.find_element(By.XPATH, "/html/body/div[2]/div/div[1]/div/form/div[3]/button[1]")
    if len(raw_string) != 4 and len(raw_string) != 3:
        btn.click()
        solve_captcha(driver)
        return
    solution = int(raw_string[:2]) + int(raw_string[-1])  
    print(f'{solution}')
    elem = driver.find_element(By.ID, "Captcha")
    elem.send_keys(solution)

    
def login(driver, PID, born_year, born_month, born_date, fail_count=0):
    if fail_count > 5:
        return -1
    driver.get(url)
    elem = driver.find_element(By.ID, "PID")
    elem.send_keys(PID)
    elem = driver.find_element(By.ID, "PBdYear")
    elem.send_keys(born_year)
    if len(born_month) == 1:
        born_month = "0" + born_month
    elem = driver.find_element(By.ID, "PBdMon")
    elem.send_keys(born_month)
    if len(born_date) == 1:
        born_date = "0" + born_date
    elem = driver.find_element(By.ID, "PBdDay")
    elem.send_keys(born_date)

    # loc = img.location
    # driver.save_screenshot('screenshot.png')
    # image = cv2.imread('screenshot.png')
    # image = image[1185: 1233, 823:968]
    # plt.imshow(image)
    # plt.show()

    # raw_string = pytesseract.image_to_string(image, lang="eng", config='--psm 13 -c tessedit_char_whitelist=0123456789').replace("\n", '')
    # print(f'"{raw_string}"')
    solve_captcha(driver)
    driver.find_element(By.ID, "login").click()


    # solution = int(raw_string[:2]) + int(raw_string[-1])  
    # print(f'{solution}')
    # input("wait")
    # elem = driver.find_element(By.ID, "Captcha")
    # elem.send_keys(solution)
    # input("wait")
    # driver.find_element(By.ID, "login").click()
    # time.sleep(0.5)
    # input("wait")
    # 要改/html/body/div[4]/div[2]/div/div/div/div/div/div/div/div[3]/div/div
    info_A = "https://ap.ceec.edu.tw/RegExam/RegInfo/RegInfoSearch?examtype=A"
    info_B = "https://ap.ceec.edu.tw/RegExam/RegInfo/RegInfoSearch?examtype=B"
    if url == url_A:
        info_url = info_A
    else:
        info_url = info_B
    if driver.current_url != info_url:
        wait = WebDriverWait(driver, 5)
        while(True):
            try:
                prompt_text = wait.until(
                    # EC.presence_of_element_located((By.CLASS_NAME, "jconfirm-title"))
                    # EC.text_to_be_present_in_element_value((By.CLASS_NAME, 'jconfirm-title'), "訊息")
                    EC.presence_of_element_located((By.XPATH, "/html/body/div[4]/div[2]/div/div/div/div/div/div/div/div[3]/div/div"))
                    # EC.visibility_of_element_located((By.XPATH, "/html/body/div[4]/div[2]/div/div/div/div/div/div/div/div[3]/div/div"))
                    ).text
            except selenium.common.exceptions.TimeoutException:
                return
            if prompt_text: break
        
        print(f'"{prompt_text}"')
        
        
        # prompt_text = driver.find_element(By.XPATH, "/html/body/div[4]/div[2]/div/div/div/div/div/div/div/div[3]/div/div").text
        # print(prompt_text)
        # prompt_text = driver.find_element(By.XPATH, "/html/body/div[4]/div[2]/div/div/div/div/div/div/div/div[3]/div/div").text
        if  prompt_text== "查無報名資料":
            print(prompt_text)
            return "no application"
        if prompt_text == "資訊填寫錯誤，請重新輸入":
            print(prompt_text)
            return "wrong info"
        if prompt_text == "應考資訊查詢已保密":
            print(prompt_text)
            return "secret"
        return login(driver, PID, born_year, born_month, born_date, fail_count+1)



def get_data(driver):
    while(True):
        ast_number = driver.find_element(By.XPATH, "/html/body/div[2]/div/div[1]/div/div[1]/div[4]").text
        if ast_number: return ast_number
    # except selenium.common.exceptions.NoSuchElementException: return -1
def set_header(sheet_obj, start, output_path):
    # objects_list = ['數學甲', '化學', '物理', '生物', '歷史', '地理', '公民', '國文', '英文', '數學Ａ', '數學Ｂ', '社會', '自然']
    if url == url_B:
        object = '分科准考證號碼'
    else: object = '學測准考證號碼'
    # for subject, i in zip(objects_list, range(13)):
    print(start)
    sheet_obj.cell(row=1, column=start).value = object
    wb_obj.save(output_path)
    
def save_data(data, sheet_obj, start, row, output_path):
    # subject_list = ["math_alpha", "chemistry", "physics", "biology", "history", "geography", "citizen", "chinese", "english", "mathA", "mathB", "society", "science"]
    # for subject, i in zip(subject_list, range(0, len(subject_list)+1)):
    #     sheet_obj.cell(row=row, column=start+i).value = data[subject]
    sheet_obj.cell(row=row, column=start).value = data
    wb_obj.save(output_path)
    
def find_start(sheet_obj, row):
    pointer = 0
    # is_empty = False
    while (True):
        pointer += 1
        val = str(sheet_obj.cell(row=row, column=pointer).value)

        if (val == 'None'): 
            if (str(sheet_obj.cell(row=row, column=pointer+1).value) == "None"): return pointer
        # if (val =='None'):
            # next_val = str(sheet_obj.cell(row=row, column=(pointer+1)).value)
            # if (next_val == "None"): 
            return pointer
def find_name(name_list, sheet_obj, row):
    pointer = 1
    while (True):
        pointer += 1
        val = str(sheet_obj.cell(row=row, column=pointer).value)
        if all([i in val for i in name_list]): return pointer
        if (val == 'None'): 
            if (str(sheet_obj.cell(row=row, column=pointer+1).value) == "None"): return -1
        # print(pointer)


if __name__ == "__main__":
    files = os.listdir('.')
    for file in files:
        if file.startswith("processing_"):
            excel_file_name = file

    # if "processing.xlsx" in files:
        # excel_file_name = "processing.xlsx"
            print(f"There is an existing processing.xlsx file whose acitve sheet is {excel_file_name}.")
            if input("Do you want to use this sheet? (y/n)") == "y":
                wb_obj = openpyxl.load_workbook(excel_file_name)
                sheet_obj = wb_obj.active
                output_path = excel_file_name
                # mode = "continue"
                break

    else:
        parser = argparse.ArgumentParser(description="This is a description of your program.")
        
        parser.add_argument("excel_file_name", help="excel原始檔")
        parser.add_argument("--sheet", help="sheet名稱")
        excel_file_name = parser.parse_args().excel_file_name
        excel_sheet_name = parser.parse_args().sheet
        wb_obj = openpyxl.load_workbook(excel_file_name)
        if excel_sheet_name:
            sheet_obj = wb_obj[excel_sheet_name]
        else: 
            sheet_obj = wb_obj.active
        excel_file_name = excel_file_name.split("/")[-1]
        output_path = f"processing_{excel_file_name}"

    # parser.add_argument("--output", required=True, help="output excel檔")
#   wait for altering
    # excels_path = "excels/黎明(方濟會)分科準0728.xlsx"
    # excels_path = "excels/0625港明分科准xlsx.xlsx"
    # excels_path = "excels/0703文華分科准.xlsx"

    
    # excels_path = "output.xlsx"

    row = sheet_obj.max_row
    column = sheet_obj.max_column
    # chrome_options = Options()
    # chrome_options.headless = True  # Set the browser to headless mode

    # Initialize the WebDriver with the specified options
    # driver = webdriver.Firefox(options=chrome_options)

    driver = webdriver.Firefox()
    if url == url_B:
        col_ast_id = find_name(["分", '准'], sheet_obj, 1)
    else:
        col_ast_id = find_name(['學', '准'], sheet_obj, 1)

    print(col_ast_id)
    if col_ast_id == -1:
        col_ast_id = find_start(sheet_obj, 1)
        print(col_ast_id)
        # input("wait")
        set_header(sheet_obj, col_ast_id, output_path)
        
    # if col_ast_id == -1:

    col_PID = find_name(["身", '證'], sheet_obj, 1)
    col_born_year = find_name(["年"], sheet_obj, 1)
    col_born_month = find_name(["月"], sheet_obj, 1)
    col_born_date = find_name(["日"], sheet_obj, 1)
    for i in range(2, row+1):
        # if mode == "continue":
        # wb_obj = openpyxl.load_workbook(output_path)
        # sheet_obj = wb_obj.active
        # if str(sheet_obj.cell(row=i, column=start).value).replace(" ", "").isnumeric(): continue
        if str(sheet_obj.cell(row=i, column=col_ast_id).value).replace(" ", "") != "None": continue

        # ID = str(sheet_obj.cell(row=i, column=col_ID).value)
        PID = str(sheet_obj.cell(row=i, column=col_PID).value)
        born_year = str(sheet_obj.cell(row=i, column=col_born_year).value)
        born_month = str(sheet_obj.cell(row=i, column=col_born_month).value)
        born_date = str(sheet_obj.cell(row=i, column=col_born_date).value)

        if len(PID) != 10: continue
        print(PID, born_year, born_month, born_date)
        success = login(driver, PID, born_year, born_month, born_date)
        # print(start)
        if success == -1:
            continue
        print("============")
        if success == "no application":
            print(success)
            save_data("未報名", sheet_obj, col_ast_id, i, output_path)
            continue
        if success == "wrong info":
            print(success)
            save_data("資料錯誤", sheet_obj, col_ast_id, i, output_path)
            continue
        if success == "secret":
            print(success)
            save_data("已保密", sheet_obj, col_ast_id, i, output_path)
            continue
        if success == -1:
            print("驗證碼錯誤次數過多")
            save_data("驗證碼錯誤次數過多", sheet_obj, col_ast_id, i, output_path)
            continue
        data = get_data(driver)
        if data == -1:
            sheet_obj.cell(row=i, column=col_ast_id).value = "登入成功但資料讀取失敗"
            continue
        # print(data)
        save_data(data, sheet_obj, col_ast_id, i, output_path)
    saved_name = output_path.split("sing_")[-1] + "(已完成分准).xlsx"
    wb_obj.save(os.path.join("finished", saved_name))
    print(saved_name)

    os.remove(output_path)

    # for root, dirs, files in os.listdir('.'):
    #     if f"processing_{excel_file_name}.xlsx" in files:

    driver.quit()
    print("finished")
